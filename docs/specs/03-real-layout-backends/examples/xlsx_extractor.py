"""ILLUSTRATIVE — proposed XlsxExtractor for app/layout.py. NOT wired in.

One `record` block per data row, fed to chunk_records (XLSX is already in
layout.TABULAR, so app/ingest.py routes it there with no router change).

Decisions (spec 03 §3 of design.md):
  - data_only=True  -> cached formula values (NOT formula strings). Caveat: a
    workbook written by openpyxl itself has NO cached values; the golden fixture
    uses literal values only (see design.md §5 fixture contract / README OQ4).
  - read_only=True  -> stream rows; a wide sheet doesn't fully materialize.
  - all sheets; first non-empty row of each sheet is its header.
  - skip fully-empty rows.
  - XLSX_MAX_ROWS cap (config, default 10000) per document; overflow dropped with
    a logged warning -- not silently exploded into embeddings.

Real seam (app/layout.py): register("xlsx", XlsxExtractor()); extract -> Document.
chunk_records consumes type=="record" blocks (per spec 02's canonical refactor).

Do NOT import this file. Port the pieces into app/layout.py.
"""
from __future__ import annotations

import datetime
import io
import logging

from openpyxl import load_workbook

from app.config import settings  # provides settings.xlsx_max_rows (see design.md §8)
from app.layout import SCHEMA_VERSION, Block, Document  # spec-02 types

log = logging.getLogger(__name__)


def _json_safe(v):
    # attrs.row is persisted via json.dumps(c.meta) in app/ingest.py, so values
    # must be JSON-serializable. openpyxl returns datetime/date/time for date
    # cells (common) -- a raw one would crash ingestion at insert (README OQ6).
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    return str(v)  # Decimal, etc. -> string, never a raw object


class XlsxExtractor:
    def extract(self, raw: str | bytes, source: str) -> Document:
        if isinstance(raw, str):
            raise TypeError("XlsxExtractor expects raw bytes, not str")
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)

        cap = settings.xlsx_max_rows
        blocks: list[Block] = []
        bid = 0
        emitted = 0
        dropped = 0

        for ws in wb.worksheets:
            header: list[str] | None = None
            for row in ws.iter_rows(values_only=True):
                if row is None or all(c is None or str(c).strip() == "" for c in row):
                    continue  # skip fully-empty rows
                if header is None:
                    header = [str(c).strip() if c is not None else f"col{n}"
                              for n, c in enumerate(row)]
                    continue
                if emitted >= cap:
                    dropped += 1
                    continue
                record = {h: _json_safe(v if v is not None else "")
                          for h, v in zip(header, row)}
                text = " | ".join(f"{h}: {v}" for h, v in record.items())
                bid += 1
                emitted += 1
                blocks.append(Block(id=f"b{bid}", type="record", text=text,
                                    attrs={"row": record, "sheet": ws.title}))

        if dropped:
            log.warning("xlsx %s: dropped %d rows over XLSX_MAX_ROWS=%d", source, dropped, cap)
        return Document(schema_version=SCHEMA_VERSION, source=source, format="xlsx",
                        meta={"sheets": wb.sheetnames}, blocks=blocks, relations=[])
